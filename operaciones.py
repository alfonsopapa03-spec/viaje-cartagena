"""
Sistema de Gestión de Proveedores
Versión 2.3 - PostgreSQL puro (BD + PDFs en BYTEA)
Contexto: Colombia
"""

import streamlit as st
import psycopg2
import psycopg2.extras
from datetime import datetime, timedelta
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.express as px

# ==================== CONFIGURACIÓN SUPABASE ====================
SUPABASE_DB_URL = "postgresql://postgres.wiomyjrmsrhcgvhgkbqe:Conejito800$@aws-1-us-west-2.pooler.supabase.com:6543/postgres"

# ==================== LISTA CIUU ====================
CIUU_OPCIONES = [
    ("", "-- Selecciona una actividad --"),
    ("0", "0 - No especificado dentro de la información"),
    ("111", "111 - Cultivo de cereales (excepto arroz), legumbres y semillas oleaginosas"),
    ("112", "112 - Cultivo de arroz"),
    ("113", "113 - Cultivo de hortalizas, raíces y tubérculos"),
    ("114", "114 - Cultivo de tabaco"),
    ("115", "115 - Cultivo de plantas textiles"),
    ("119", "119 - Otros cultivos transitorios n.c.p."),
    ("121", "121 - Cultivo de frutas tropicales y subtropicales"),
    ("122", "122 - Cultivo de plátano y banano"),
    ("123", "123 - Cultivo de café"),
    ("124", "124 - Cultivo de caña de azúcar"),
    ("125", "125 - Cultivo de flor de corte"),
    ("126", "126 - Cultivo de palma para aceite (palma africana) y otros frutos oleaginosos"),
    ("127", "127 - Cultivo de plantas con las que se preparan bebidas"),
    ("128", "128 - Cultivo de especias y de plantas aromáticas y medicinales"),
    ("129", "129 - Otros cultivos permanentes n.c.p."),
    ("130", "130 - Propagación de plantas (actividades de los viveros, excepto viveros forestales)"),
    ("141", "141 - Cría de ganado bovino y bufalino"),
    ("142", "142 - Cría de caballos y otros equinos"),
    ("143", "143 - Cría de ovejas y cabras"),
    ("144", "144 - Cría de ganado porcino"),
    ("145", "145 - Cría de aves de corral"),
    ("149", "149 - Cría de otros animales n.c.p."),
    ("150", "150 - Explotación mixta (agrícola y pecuaria)"),
    ("161", "161 - Actividades de apoyo a la agricultura"),
    ("162", "162 - Actividades de apoyo a la ganadería"),
    ("163", "163 - Actividades posteriores a la cosecha"),
    ("164", "164 - Tratamiento de semillas para propagación"),
    ("170", "170 - Caza ordinaria y mediante trampas y actividades de servicios conexas"),
    ("210", "210 - Silvicultura y otras actividades forestales"),
    ("220", "220 - Extracción de madera"),
    ("230", "230 - Recolección de productos forestales diferentes a la madera"),
    ("240", "240 - Servicios de apoyo a la silvicultura"),
    ("311", "311 - Pesca marítima"),
    ("312", "312 - Pesca de agua dulce"),
    ("321", "321 - Acuicultura marítima"),
    ("322", "322 - Acuicultura de agua dulce"),
    ("510", "510 - Extracción de hulla (carbón de piedra)"),
    ("520", "520 - Extracción de carbón lignito"),
    ("610", "610 - Extracción de petróleo crudo"),
    ("620", "620 - Extracción de gas natural"),
    ("710", "710 - Extracción de minerales de hierro"),
    ("721", "721 - Extracción de minerales de uranio y de torio"),
    ("722", "722 - Extracción de oro y otros metales preciosos"),
    ("723", "723 - Extracción de minerales de níquel"),
    ("729", "729 - Extracción de otros minerales metalíferos no ferrosos n.c.p."),
    ("811", "811 - Extracción de piedra, arena, arcillas comunes, yeso y anhidrita"),
    ("812", "812 - Extracción de arcillas de uso industrial, Caliza, Caolín y Bentonitas"),
    ("820", "820 - Extracción de esmeraldas, piedras preciosas y semipreciosas"),
    ("891", "891 - Extracción de minerales para la fabricación de abonos y productos químicos"),
    ("892", "892 - Extracción de halita (sal)"),
    ("899", "899 - Extracción de otros minerales no metálicos n.c.p."),
    ("910", "910 - Actividades de apoyo para la extracción de petróleo y de gas natural"),
    ("990", "990 - Actividades de apoyo para otras actividades de explotación de minas y canteras"),
    ("1011", "1011 - Procesamiento y conservación de carne y productos cárnicos"),
    ("1012", "1012 - Procesamiento y conservación de pescados, crustáceos y moluscos"),
    ("1020", "1020 - Procesamiento y conservación de frutas, legumbres, hortalizas y tubérculos"),
    ("1030", "1030 - Elaboración de aceites y grasas de origen vegetal y animal"),
    ("1040", "1040 - Elaboración de productos lácteos"),
    ("1051", "1051 - Elaboración de productos de molinería"),
    ("1052", "1052 - Elaboración de almidones y productos derivados del almidón"),
    ("1061", "1061 - Trilla de café"),
    ("1062", "1062 - Descafeinado, tostión y molienda del café"),
    ("1063", "1063 - Otros derivados del café"),
    ("1071", "1071 - Elaboración y refinación de azúcar"),
    ("1072", "1072 - Elaboración de panela"),
    ("1081", "1081 - Elaboración de productos de panadería"),
    ("1082", "1082 - Elaboración de cacao, chocolate y productos de confitería"),
    ("1083", "1083 - Elaboración de macarrones, fideos, alcuzcuz y productos farináceos similares"),
    ("1084", "1084 - Elaboración de comidas y platos preparados"),
    ("1089", "1089 - Elaboración de otros productos alimenticios n.c.p."),
    ("1090", "1090 - Elaboración de alimentos preparados para animales"),
    ("1101", "1101 - Destilación, rectificación y mezcla de bebidas alcohólicas"),
    ("1102", "1102 - Elaboración de bebidas fermentadas no destiladas"),
    ("1103", "1103 - Producción de malta, elaboración de cervezas y otras bebidas malteadas"),
    ("1104", "1104 - Elaboración de bebidas no alcohólicas, producción de aguas minerales y embotelladas"),
    ("1200", "1200 - Elaboración de productos de tabaco"),
    ("1311", "1311 - Preparación e Hilatura de fibras textiles"),
    ("1312", "1312 - Tejeduría de productos textiles"),
    ("1313", "1313 - Acabado de productos textiles"),
    ("1391", "1391 - Fabricación de tejidos de punto y ganchillo"),
    ("1392", "1392 - Confección de artículos con materiales textiles, excepto prendas de vestir"),
    ("1393", "1393 - Fabricación de tapetes y alfombras para pisos"),
    ("1394", "1394 - Fabricación de cuerdas, cordeles, cables, bramantes y redes"),
    ("1399", "1399 - Fabricación de otros artículos textiles n.c.p."),
    ("1410", "1410 - Confección de prendas de vestir, excepto prendas de piel"),
    ("1420", "1420 - Fabricación de artículos de piel"),
    ("1430", "1430 - Fabricación de artículos de punto y ganchillo"),
    ("1511", "1511 - Curtido y recurtido de cueros; recurtido y teñido de pieles"),
    ("1512", "1512 - Fabricación de artículos de viaje, bolsos de mano y artículos de talabartería (cuero)"),
    ("1513", "1513 - Fabricación de artículos de viaje, bolsos de mano y artículos de talabartería (otros materiales)"),
    ("1521", "1521 - Fabricación de calzado de cuero y piel, con cualquier tipo de suela"),
    ("1522", "1522 - Fabricación de otros tipos de calzado, excepto calzado de cuero y piel"),
    ("1523", "1523 - Fabricación de partes del calzado"),
    ("1610", "1610 - Aserrado, acepillado e impregnación de la madera"),
    ("1620", "1620 - Fabricación de hojas de madera para enchapado; tableros contrachapados y paneles"),
    ("1630", "1630 - Fabricación de partes y piezas de madera, carpintería y ebanistería para construcción"),
    ("1640", "1640 - Fabricación de recipientes de madera"),
    ("1690", "1690 - Fabricación de otros productos de madera; artículos de corcho, cestería y espartería"),
    ("1701", "1701 - Fabricación de pulpas (pastas) celulósicas; papel y cartón"),
    ("1702", "1702 - Fabricación de papel y cartón ondulado; envases, empaques y embalajes de papel y cartón"),
    ("1709", "1709 - Fabricación de otros artículos de papel y cartón"),
    ("1811", "1811 - Actividades de impresión"),
    ("1812", "1812 - Actividades de servicios relacionados con la impresión"),
    ("1820", "1820 - Producción de copias a partir de grabaciones originales"),
    ("1910", "1910 - Fabricación de productos de hornos de coque"),
    ("1921", "1921 - Fabricación de productos de la refinación del petróleo"),
    ("2011", "2011 - Fabricación de sustancias y productos químicos básicos"),
    ("2012", "2012 - Fabricación de abonos y compuestos inorgánicos nitrogenados"),
    ("2013", "2013 - Fabricación de plásticos en formas primarias"),
    ("2014", "2014 - Fabricación de caucho sintético en formas primarias"),
    ("2021", "2021 - Fabricación de plaguicidas y otros productos químicos de uso agropecuario"),
    ("2022", "2022 - Fabricación de pinturas, barnices, tintas para impresión y masillas"),
    ("2023", "2023 - Fabricación de jabones, detergentes, perfumes y preparados de tocador"),
    ("2029", "2029 - Fabricación de otros productos químicos n.c.p."),
    ("2030", "2030 - Fabricación de fibras sintéticas y artificiales"),
    ("2100", "2100 - Fabricación de productos farmacéuticos y productos botánicos de uso farmacéutico"),
    ("2211", "2211 - Fabricación de llantas y neumáticos de caucho"),
    ("2212", "2212 - Reencauche de llantas usadas"),
    ("2219", "2219 - Fabricación de formas básicas de caucho y otros productos de caucho n.c.p."),
    ("2221", "2221 - Fabricación de formas básicas de plástico"),
    ("2229", "2229 - Fabricación de artículos de plástico n.c.p."),
    ("2310", "2310 - Fabricación de vidrio y productos de vidrio"),
    ("2391", "2391 - Fabricación de productos refractarios"),
    ("2392", "2392 - Fabricación de materiales de arcilla para la construcción"),
    ("2393", "2393 - Fabricación de otros productos de cerámica y porcelana"),
    ("2394", "2394 - Fabricación de cemento, cal y yeso"),
    ("2395", "2395 - Fabricación de artículos de hormigón, cemento y yeso"),
    ("2396", "2396 - Corte, tallado y acabado de la piedra"),
    ("2399", "2399 - Fabricación de otros productos minerales no metálicos n.c.p."),
    ("2410", "2410 - Industrias básicas de hierro y de acero"),
    ("2421", "2421 - Industrias básicas de metales preciosos"),
    ("2429", "2429 - Industrias básicas de otros metales no ferrosos"),
    ("2431", "2431 - Fundición de hierro y de acero"),
    ("2432", "2432 - Fundición de metales no ferrosos"),
    ("2511", "2511 - Fabricación de productos metálicos para uso estructural"),
    ("2512", "2512 - Fabricación de tanques, depósitos y recipientes de metal"),
    ("2513", "2513 - Fabricación de generadores de vapor"),
    ("2520", "2520 - Fabricación de armas y municiones"),
    ("2591", "2591 - Forja, prensado, estampado y laminado de metal; pulvimetalurgia"),
    ("2592", "2592 - Tratamiento y revestimiento de metales; mecanizado"),
    ("2593", "2593 - Fabricación de artículos de cuchillería, herramientas de mano y ferretería"),
    ("2599", "2599 - Fabricación de otros productos elaborados de metal n.c.p."),
    ("2610", "2610 - Fabricación de componentes y tableros electrónicos"),
    ("2620", "2620 - Fabricación de computadoras y de equipo periférico"),
    ("2630", "2630 - Fabricación de equipos de comunicación"),
    ("2640", "2640 - Fabricación de aparatos electrónicos de consumo"),
    ("2651", "2651 - Fabricación de equipo de medición, prueba, navegación y control"),
    ("2652", "2652 - Fabricación de relojes"),
    ("2660", "2660 - Fabricación de equipo de irradiación y equipo electrónico de uso médico"),
    ("2670", "2670 - Fabricación de instrumentos ópticos y equipo fotográfico"),
    ("2680", "2680 - Fabricación de medios magnéticos y ópticos para almacenamiento de datos"),
    ("2711", "2711 - Fabricación de motores, generadores y transformadores eléctricos"),
    ("2712", "2712 - Fabricación de aparatos de distribución y control de la energía eléctrica"),
    ("2720", "2720 - Fabricación de pilas, baterías y acumuladores eléctricos"),
    ("2731", "2731 - Fabricación de hilos y cables eléctricos y de fibra óptica"),
    ("2732", "2732 - Fabricación de dispositivos de cableado"),
    ("2740", "2740 - Fabricación de equipos eléctricos de iluminación"),
    ("2750", "2750 - Fabricación de aparatos de uso doméstico"),
    ("2790", "2790 - Fabricación de otros tipos de equipo eléctrico n.c.p."),
    ("2811", "2811 - Fabricación de motores, turbinas y partes para motores de combustión interna"),
    ("2812", "2812 - Fabricación de equipos de potencia hidráulica y neumática"),
    ("2813", "2813 - Fabricación de otras bombas, compresores, grifos y válvulas"),
    ("2814", "2814 - Fabricación de cojinetes, engranajes, trenes de engranajes y piezas de transmisión"),
    ("2815", "2815 - Fabricación de hornos, hogares y quemadores industriales"),
    ("2816", "2816 - Fabricación de equipo de elevación y manipulación"),
    ("2817", "2817 - Fabricación de maquinaria y equipo de oficina (excepto computadoras)"),
    ("2818", "2818 - Fabricación de herramientas manuales con motor"),
    ("2819", "2819 - Fabricación de otros tipos de maquinaria y equipo de uso general n.c.p."),
    ("2821", "2821 - Fabricación de maquinaria agropecuaria y forestal"),
    ("2822", "2822 - Fabricación de máquinas formadoras de metal y de máquinas herramienta"),
    ("2823", "2823 - Fabricación de maquinaria para la metalurgia"),
    ("2824", "2824 - Fabricación de maquinaria para explotación de minas, canteras y construcción"),
    ("2825", "2825 - Fabricación de maquinaria para la elaboración de alimentos, bebidas y tabaco"),
    ("2826", "2826 - Fabricación de maquinaria para elaboración de textiles, prendas de vestir y cueros"),
    ("2829", "2829 - Fabricación de otros tipos de maquinaria y equipo de uso especial n.c.p."),
    ("2910", "2910 - Fabricación de vehículos automotores y sus motores"),
    ("2920", "2920 - Fabricación de carrocerías para vehículos automotores; remolques y semirremolques"),
    ("2930", "2930 - Fabricación de partes, piezas (autopartes) y accesorios para vehículos automotores"),
    ("3011", "3011 - Construcción de barcos y de estructuras flotantes"),
    ("3012", "3012 - Construcción de embarcaciones de recreo y deporte"),
    ("3020", "3020 - Fabricación de locomotoras y de material rodante para ferrocarriles"),
    ("3030", "3030 - Fabricación de aeronaves, naves espaciales y de maquinaria conexa"),
    ("3040", "3040 - Fabricación de vehículos militares de combate"),
    ("3091", "3091 - Fabricación de motocicletas"),
    ("3092", "3092 - Fabricación de bicicletas y de sillas de ruedas para personas con discapacidad"),
    ("3099", "3099 - Fabricación de otros tipos de equipo de transporte n.c.p."),
    ("3110", "3110 - Fabricación de muebles"),
    ("3120", "3120 - Fabricación de colchones y somieres"),
    ("3210", "3210 - Fabricación de joyas, bisutería y artículos conexos"),
    ("3220", "3220 - Fabricación de instrumentos musicales"),
    ("3230", "3230 - Fabricación de artículos y equipo para la práctica del deporte"),
    ("3240", "3240 - Fabricación de juegos, juguetes y rompecabezas"),
    ("3250", "3250 - Fabricación de instrumentos, aparatos y materiales médicos y odontológicos"),
    ("3290", "3290 - Otras industrias manufactureras n.c.p."),
    ("3311", "3311 - Mantenimiento y reparación especializado de productos elaborados en metal"),
    ("3312", "3312 - Mantenimiento y reparación especializado de maquinaria y equipo"),
    ("3313", "3313 - Mantenimiento y reparación especializado de equipo electrónico y óptico"),
    ("3314", "3314 - Mantenimiento y reparación especializado de equipo eléctrico"),
    ("3315", "3315 - Mantenimiento y reparación especializado de equipo de transporte (excepto automotores)"),
    ("3319", "3319 - Mantenimiento y reparación de otros tipos de equipos y sus componentes n.c.p."),
    ("3320", "3320 - Instalación especializada de maquinaria y equipo industrial"),
    ("3511", "3511 - Generación de energía eléctrica"),
    ("3512", "3512 - Transmisión de energía eléctrica"),
    ("3513", "3513 - Distribución de energía eléctrica"),
    ("3514", "3514 - Comercialización de energía eléctrica"),
    ("3520", "3520 - Producción de gas; distribución de combustibles gaseosos por tuberías"),
    ("3530", "3530 - Suministro de vapor y aire acondicionado"),
    ("3600", "3600 - Captación, tratamiento y distribución de agua"),
    ("3700", "3700 - Evacuación y tratamiento de aguas residuales"),
    ("3811", "3811 - Recolección de desechos no peligrosos"),
    ("3812", "3812 - Recolección de desechos peligrosos"),
    ("3821", "3821 - Tratamiento y disposición de desechos no peligrosos"),
    ("3822", "3822 - Tratamiento y disposición de desechos peligrosos"),
    ("3830", "3830 - Recuperación de materiales"),
    ("3900", "3900 - Actividades de saneamiento ambiental y otros servicios de gestión de desechos"),
    ("4111", "4111 - Construcción de edificios residenciales"),
    ("4112", "4112 - Construcción de edificios no residenciales"),
    ("4210", "4210 - Construcción de carreteras y vías de ferrocarril"),
    ("4220", "4220 - Construcción de proyectos de servicio público"),
    ("4290", "4290 - Construcción de otras obras de ingeniería civil"),
    ("4311", "4311 - Demolición"),
    ("4312", "4312 - Preparación del terreno"),
    ("4321", "4321 - Instalaciones eléctricas"),
    ("4322", "4322 - Instalaciones de fontanería, calefacción y aire acondicionado"),
    ("4329", "4329 - Otras instalaciones especializadas"),
    ("4330", "4330 - Terminación y acabado de edificios y obras de ingeniería civil"),
    ("4390", "4390 - Otras actividades especializadas para la construcción de edificios y obras de ingeniería civil"),
    ("4511", "4511 - Comercio de vehículos automotores nuevos"),
    ("4512", "4512 - Comercio de vehículos automotores usados"),
    ("4520", "4520 - Mantenimiento y reparación de vehículos automotores"),
    ("4530", "4530 - Comercio de partes, piezas (autopartes) y accesorios para vehículos automotores"),
    ("4541", "4541 - Comercio de motocicletas y de sus partes, piezas y accesorios"),
    ("4542", "4542 - Mantenimiento y reparación de motocicletas y de sus partes y piezas"),
    ("4610", "4610 - Comercio al por mayor a cambio de una retribución o por contrata"),
    ("4620", "4620 - Comercio al por mayor de materias primas agropecuarias; animales vivos"),
    ("4631", "4631 - Comercio al por mayor de productos alimenticios"),
    ("4632", "4632 - Comercio al por mayor de bebidas y tabaco"),
    ("4641", "4641 - Comercio al por mayor de productos textiles; productos confeccionados para uso doméstico"),
    ("4642", "4642 - Comercio al por mayor de prendas de vestir"),
    ("4643", "4643 - Comercio al por mayor de calzado"),
    ("4644", "4644 - Comercio al por mayor de aparatos y equipo de uso doméstico"),
    ("4645", "4645 - Comercio al por mayor de productos farmacéuticos, medicinales, cosméticos y de tocador"),
    ("4649", "4649 - Comercio al por mayor de otros utensilios domésticos n.c.p."),
    ("4651", "4651 - Comercio al por mayor de computadores, equipo periférico y programas de informática"),
    ("4652", "4652 - Comercio al por mayor de equipo, partes y piezas electrónicos y de telecomunicaciones"),
    ("4653", "4653 - Comercio al por mayor de maquinaria y equipo agropecuarios"),
    ("4659", "4659 - Comercio al por mayor de otros tipos de maquinaria y equipo n.c.p."),
    ("4661", "4661 - Comercio al por mayor de combustibles sólidos, líquidos, gaseosos y productos conexos"),
    ("4662", "4662 - Comercio al por mayor de metales y productos metalíferos"),
    ("4663", "4663 - Comercio al por mayor de materiales de construcción, ferretería, pinturas y vidrio"),
    ("4664", "4664 - Comercio al por mayor de productos químicos básicos, cauchos y plásticos"),
    ("4665", "4665 - Comercio al por mayor de desperdicios, desechos y chatarra"),
    ("4669", "4669 - Comercio al por mayor de otros productos n.c.p."),
    ("4690", "4690 - Comercio al por mayor no especializado"),
    ("4711", "4711 - Comercio al por menor en establecimientos no especializados (alimentos, bebidas o tabaco)"),
    ("4719", "4719 - Comercio al por menor en establecimientos no especializados (otros productos)"),
    ("4721", "4721 - Comercio al por menor de productos agrícolas para el consumo"),
    ("4722", "4722 - Comercio al por menor de leche, productos lácteos y huevos"),
    ("4723", "4723 - Comercio al por menor de carnes, productos cárnicos, pescados y productos de mar"),
    ("4724", "4724 - Comercio al por menor de bebidas y productos de tabaco"),
    ("4729", "4729 - Comercio al por menor de otros productos alimenticios n.c.p."),
    ("4731", "4731 - Comercio al por menor de combustible para automotores"),
    ("4732", "4732 - Comercio al por menor de lubricantes, aditivos y productos de limpieza para vehículos"),
    ("4741", "4741 - Comercio al por menor de computadores, equipos periféricos y telecomunicaciones"),
    ("4742", "4742 - Comercio al por menor de equipos y aparatos de sonido y de video"),
    ("4751", "4751 - Comercio al por menor de productos textiles"),
    ("4752", "4752 - Comercio al por menor de artículos de ferretería, pinturas y productos de vidrio"),
    ("4753", "4753 - Comercio al por menor de tapices, alfombras y recubrimientos para paredes y pisos"),
    ("4754", "4754 - Comercio al por menor de electrodomésticos, muebles y equipos de iluminación"),
    ("4755", "4755 - Comercio al por menor de artículos y utensilios de uso doméstico"),
    ("4759", "4759 - Comercio al por menor de otros artículos domésticos"),
    ("4761", "4761 - Comercio al por menor de libros, periódicos, papelería y escritorio"),
    ("4762", "4762 - Comercio al por menor de artículos deportivos"),
    ("4769", "4769 - Comercio al por menor de otros artículos culturales y de entretenimiento n.c.p."),
    ("4771", "4771 - Comercio al por menor de prendas de vestir y sus accesorios"),
    ("4772", "4772 - Comercio al por menor de calzado y artículos de cuero"),
    ("4773", "4773 - Comercio al por menor de productos farmacéuticos, cosméticos y artículos de tocador"),
    ("4774", "4774 - Comercio al por menor de otros productos nuevos"),
    ("4775", "4775 - Comercio al por menor de artículos de segunda mano"),
    ("4781", "4781 - Comercio al por menor de alimentos, bebidas y tabaco en puestos de venta móviles"),
    ("4782", "4782 - Comercio al por menor de textiles, prendas de vestir y calzado en puestos móviles"),
    ("4789", "4789 - Comercio al por menor de otros productos en puestos de venta móviles"),
    ("4791", "4791 - Comercio al por menor realizado a través de Internet"),
    ("4792", "4792 - Comercio al por menor realizado a través de casas de venta o por correo"),
    ("4799", "4799 - Otros tipos de comercio al por menor no realizado en establecimientos"),
    ("4911", "4911 - Transporte férreo de pasajeros"),
    ("4912", "4912 - Transporte férreo de carga"),
    ("4921", "4921 - Transporte de pasajeros"),
    ("4922", "4922 - Transporte mixto"),
    ("4923", "4923 - Transporte de carga por carretera"),
    ("4930", "4930 - Transporte por tuberías"),
    ("5011", "5011 - Transporte de pasajeros marítimo y de cabotaje"),
    ("5012", "5012 - Transporte de carga marítimo y de cabotaje"),
    ("5021", "5021 - Transporte fluvial de pasajeros"),
    ("5022", "5022 - Transporte fluvial de carga"),
    ("5111", "5111 - Transporte aéreo nacional de pasajeros"),
    ("5112", "5112 - Transporte aéreo internacional de pasajeros"),
    ("5121", "5121 - Transporte aéreo nacional de carga"),
    ("5122", "5122 - Transporte aéreo internacional de carga"),
    ("5210", "5210 - Almacenamiento y depósito"),
    ("5221", "5221 - Actividades de estaciones, vías y servicios complementarios para el transporte terrestre"),
    ("5222", "5222 - Actividades de puertos y servicios complementarios para el transporte acuático"),
    ("5223", "5223 - Actividades de aeropuertos, servicios de navegación aérea y conexos"),
    ("5224", "5224 - Manipulación de carga"),
    ("5229", "5229 - Otras actividades complementarias al transporte"),
    ("5310", "5310 - Actividades postales nacionales"),
    ("5320", "5320 - Actividades de mensajería"),
    ("5511", "5511 - Alojamiento en hoteles"),
    ("5512", "5512 - Alojamiento en aparta-hoteles"),
    ("5513", "5513 - Alojamiento en centros vacacionales"),
    ("5514", "5514 - Alojamiento rural"),
    ("5519", "5519 - Otros tipos de alojamientos para visitantes"),
    ("5520", "5520 - Actividades de zonas de camping y parques para vehículos recreacionales"),
    ("5530", "5530 - Servicios por horas"),
    ("5590", "5590 - Otros tipos de alojamiento n.c.p."),
    ("5611", "5611 - Expendio a la mesa de comidas preparadas"),
    ("5612", "5612 - Expendio por autoservicio de comidas preparadas"),
    ("5613", "5613 - Expendio de comidas preparadas en cafeterías"),
    ("5619", "5619 - Otros tipos de expendio de comidas preparadas n.c.p."),
    ("5621", "5621 - Catering para eventos"),
    ("5629", "5629 - Actividades de otros servicios de comidas"),
    ("5630", "5630 - Expendio de bebidas alcohólicas para el consumo dentro del establecimiento"),
    ("5811", "5811 - Edición de libros"),
    ("5812", "5812 - Edición de directorios y listas de correo"),
    ("5813", "5813 - Edición de periódicos, revistas y otras publicaciones periódicas"),
    ("5819", "5819 - Otros trabajos de edición"),
    ("5820", "5820 - Edición de programas de informática (software)"),
    ("5911", "5911 - Actividades de producción de películas cinematográficas, videos y comerciales de televisión"),
    ("5912", "5912 - Actividades de postproducción de películas cinematográficas y videos"),
    ("5913", "5913 - Actividades de distribución de películas cinematográficas y videos"),
    ("5914", "5914 - Actividades de exhibición de películas cinematográficas y videos"),
    ("5920", "5920 - Actividades de grabación de sonido y edición de música"),
    ("6010", "6010 - Actividades de programación y transmisión en el servicio de radiodifusión sonora"),
    ("6020", "6020 - Actividades de programación y transmisión de televisión"),
    ("6110", "6110 - Actividades de telecomunicaciones alámbricas"),
    ("6120", "6120 - Actividades de telecomunicaciones inalámbricas"),
    ("6130", "6130 - Actividades de telecomunicación satelital"),
    ("6190", "6190 - Otras actividades de telecomunicaciones"),
    ("6201", "6201 - Actividades de desarrollo de sistemas informáticos"),
    ("6202", "6202 - Actividades de consultoría informática y administración de instalaciones informáticas"),
    ("6209", "6209 - Otras actividades de tecnologías de información y servicios informáticos"),
    ("6311", "6311 - Procesamiento de datos, alojamiento (hosting) y actividades relacionadas"),
    ("6312", "6312 - Portales Web"),
    ("6391", "6391 - Actividades de agencias de noticias"),
    ("6399", "6399 - Otras actividades de servicio de información n.c.p."),
    ("6411", "6411 - Banco Central"),
    ("6412", "6412 - Bancos comerciales"),
    ("6421", "6421 - Actividades de las corporaciones financieras"),
    ("6422", "6422 - Actividades de las compañías de financiamiento"),
    ("6423", "6423 - Banca de segundo piso"),
    ("6424", "6424 - Actividades de las cooperativas financieras"),
    ("6431", "6431 - Fideicomisos, fondos y entidades financieras similares"),
    ("6432", "6432 - Fondos de cesantías"),
    ("6491", "6491 - Leasing financiero (arrendamiento financiero)"),
    ("6492", "6492 - Actividades financieras de fondos de empleados y formas asociativas del sector solidario"),
    ("6493", "6493 - Actividades de compra de cartera o factoring"),
    ("6494", "6494 - Otras actividades de distribución de fondos"),
    ("6495", "6495 - Instituciones especiales oficiales"),
    ("6499", "6499 - Otras actividades de servicio financiero n.c.p."),
    ("6511", "6511 - Seguros generales"),
    ("6512", "6512 - Seguros de vida"),
    ("6513", "6513 - Reaseguros"),
    ("6514", "6514 - Capitalización"),
    ("6521", "6521 - Servicios de seguros sociales de salud"),
    ("6522", "6522 - Servicios de seguros sociales de riesgos profesionales"),
    ("6531", "6531 - Régimen de prima media con prestación definida (RPM)"),
    ("6532", "6532 - Régimen de ahorro individual (RAI)"),
    ("6611", "6611 - Administración de mercados financieros"),
    ("6612", "6612 - Corretaje de valores y de contratos de productos básicos"),
    ("6613", "6613 - Otras actividades relacionadas con el mercado de valores"),
    ("6614", "6614 - Actividades de las casas de cambio"),
    ("6615", "6615 - Actividades de los profesionales de compra y venta de divisas"),
    ("6619", "6619 - Otras actividades auxiliares de las actividades de servicios financieros n.c.p."),
    ("6621", "6621 - Actividades de agentes y corredores de seguros"),
    ("6629", "6629 - Evaluación de riesgos y daños y otras actividades de servicios auxiliares"),
    ("6630", "6630 - Actividades de administración de fondos"),
    ("6810", "6810 - Actividades inmobiliarias realizadas con bienes propios o arrendados"),
    ("6820", "6820 - Actividades inmobiliarias realizadas a cambio de una retribución o por contrata"),
    ("6910", "6910 - Actividades jurídicas"),
    ("6920", "6920 - Actividades de contabilidad, teneduría de libros, auditoria financiera y asesoría tributaria"),
    ("7010", "7010 - Actividades de administración empresarial"),
    ("7020", "7020 - Actividades de consultoría de gestión"),
    ("7110", "7110 - Actividades de arquitectura e ingeniería y consultoría técnica"),
    ("7120", "7120 - Ensayos y análisis técnicos"),
    ("7210", "7210 - Investigaciones y desarrollo experimental en ciencias naturales e ingeniería"),
    ("7220", "7220 - Investigaciones y desarrollo experimental en ciencias sociales y humanidades"),
    ("7310", "7310 - Publicidad"),
    ("7320", "7320 - Estudios de mercado y realización de encuestas de opinión pública"),
    ("7410", "7410 - Actividades especializadas de diseño"),
    ("7420", "7420 - Actividades de fotografía"),
    ("7490", "7490 - Otras actividades profesionales, científicas y técnicas n.c.p."),
    ("7500", "7500 - Actividades veterinarias"),
    ("7710", "7710 - Alquiler y arrendamiento de vehículos automotores"),
    ("7721", "7721 - Alquiler y arrendamiento de equipo recreativo y deportivo"),
    ("7722", "7722 - Alquiler de videos y discos"),
    ("7729", "7729 - Alquiler y arrendamiento de otros efectos personales y enseres domésticos n.c.p."),
    ("7730", "7730 - Alquiler y arrendamiento de otros tipos de maquinaria, equipo y bienes tangibles n.c.p."),
    ("7740", "7740 - Arrendamiento de propiedad intelectual y productos similares"),
    ("7810", "7810 - Actividades de agencias de empleo"),
    ("7820", "7820 - Actividades de agencias de empleo temporal"),
    ("7830", "7830 - Otras actividades de suministro de recurso humano"),
    ("7911", "7911 - Actividades de las agencias de viajes"),
    ("7912", "7912 - Actividades de operadores turísticos"),
    ("7990", "7990 - Otros servicios de reserva y actividades relacionadas"),
    ("8010", "8010 - Actividades de seguridad privada"),
    ("8020", "8020 - Actividades de servicios de sistemas de seguridad"),
    ("8030", "8030 - Actividades de detectives e investigadores privados"),
    ("8110", "8110 - Actividades combinadas de apoyo a instalaciones"),
    ("8121", "8121 - Limpieza general interior de edificios"),
    ("8129", "8129 - Otras actividades de limpieza de edificios e instalaciones industriales"),
    ("8130", "8130 - Actividades de paisajismo y servicios de mantenimiento conexos"),
    ("8211", "8211 - Actividades combinadas de servicios administrativos de oficina"),
    ("8219", "8219 - Fotocopiado, preparación de documentos y otras actividades especializadas de apoyo a oficina"),
    ("8220", "8220 - Actividades de centros de llamadas (Call center)"),
    ("8230", "8230 - Organización de convenciones y eventos comerciales"),
    ("8291", "8291 - Actividades de agencias de cobranza y oficinas de calificación crediticia"),
    ("8292", "8292 - Actividades de envase y empaque"),
    ("8299", "8299 - Otras actividades de servicio de apoyo a las empresas n.c.p."),
    ("8411", "8411 - Actividades legislativas de la administración publica"),
    ("8412", "8412 - Actividades ejecutivas de la administración publica"),
    ("8413", "8413 - Regulación de organismos que prestan servicios de salud, educativos, culturales y sociales"),
    ("8414", "8414 - Actividades reguladoras y facilitadoras de la actividad económica"),
    ("8415", "8415 - Actividades de los otros órganos de control"),
    ("8421", "8421 - Relaciones exteriores"),
    ("8422", "8422 - Actividades de defensa"),
    ("8423", "8423 - Orden público y actividades de seguridad"),
    ("8424", "8424 - Administración de justicia"),
    ("8430", "8430 - Actividades de planes de seguridad social de afiliación obligatoria"),
    ("8512", "8512 - Educación preescolar"),
    ("8513", "8513 - Educación básica primaria"),
    ("8521", "8521 - Educación básica secundaria"),
    ("8522", "8522 - Educación media académica"),
    ("8523", "8523 - Educación media técnica y de formación laboral"),
    ("8530", "8530 - Establecimientos que combinan diferentes niveles de educación"),
    ("8541", "8541 - Educación técnica profesional"),
    ("8542", "8542 - Educación tecnológica"),
    ("8543", "8543 - Educación de instituciones universitarias o de escuelas tecnológicas"),
    ("8544", "8544 - Educación de universidades"),
    ("8551", "8551 - Formación académica no formal"),
    ("8552", "8552 - Enseñanza deportiva y recreativa"),
    ("8553", "8553 - Enseñanza cultural"),
    ("8559", "8559 - Otros tipos de educación n.c.p."),
    ("8560", "8560 - Actividades de apoyo a la educación"),
    ("8610", "8610 - Actividades de hospitales y clínicas, con internación"),
    ("8621", "8621 - Actividades de la práctica médica, sin internación"),
    ("8622", "8622 - Actividades de la práctica odontológica"),
    ("8691", "8691 - Actividades de apoyo diagnóstico"),
    ("8692", "8692 - Actividades de apoyo terapéutico"),
    ("8699", "8699 - Otras actividades de atención de la salud humana"),
    ("8710", "8710 - Actividades de atención residencial medicalizada de tipo general"),
    ("8720", "8720 - Actividades de atención residencial para pacientes con retardo mental y consumo de SPA"),
    ("8730", "8730 - Actividades de atención en instituciones para personas mayores y/o discapacitadas"),
    ("8790", "8790 - Otras actividades de atención en instituciones con alojamiento"),
    ("8810", "8810 - Actividades de asistencia social sin alojamiento para personas mayores y discapacitadas"),
    ("8890", "8890 - Otras actividades de asistencia social sin alojamiento"),
    ("9001", "9001 - Creación literaria"),
    ("9002", "9002 - Creación musical"),
    ("9003", "9003 - Creación teatral"),
    ("9004", "9004 - Creación audiovisual"),
    ("9005", "9005 - Artes plásticas y visuales"),
    ("9006", "9006 - Actividades teatrales"),
    ("9007", "9007 - Actividades de espectáculos musicales en vivo"),
    ("9008", "9008 - Otras actividades de espectáculos en vivo"),
    ("9101", "9101 - Actividades de bibliotecas y archivos"),
    ("9102", "9102 - Actividades y funcionamiento de museos, conservación de edificios y sitios históricos"),
    ("9103", "9103 - Actividades de jardines botánicos, zoológicos y reservas naturales"),
    ("9200", "9200 - Actividades de juegos de azar y apuestas"),
    ("9311", "9311 - Gestión de instalaciones deportivas"),
    ("9312", "9312 - Actividades de clubes deportivos"),
    ("9319", "9319 - Otras actividades deportivas"),
    ("9321", "9321 - Actividades de parques de atracciones y parques temáticos"),
    ("9329", "9329 - Otras actividades recreativas y de esparcimiento n.c.p."),
    ("9411", "9411 - Actividades de asociaciones empresariales y de empleadores"),
    ("9412", "9412 - Actividades de asociaciones profesionales"),
    ("9420", "9420 - Actividades de sindicatos de empleados"),
    ("9491", "9491 - Actividades de asociaciones religiosas"),
    ("9492", "9492 - Actividades de asociaciones políticas"),
    ("9499", "9499 - Actividades de otras asociaciones n.c.p."),
    ("9511", "9511 - Mantenimiento y reparación de computadoras y de equipo periférico"),
    ("9512", "9512 - Mantenimiento y reparación de equipos de comunicación"),
    ("9521", "9521 - Mantenimiento y reparación de aparatos electrónicos de consumo"),
    ("9522", "9522 - Mantenimiento y reparación de aparatos y equipos domésticos y de jardinería"),
    ("9523", "9523 - Reparación de calzado y artículos de cuero"),
    ("9524", "9524 - Reparación de muebles y accesorios para el hogar"),
    ("9529", "9529 - Mantenimiento y reparación de otros efectos personales y enseres domésticos"),
    ("9601", "9601 - Lavado y limpieza, incluso la limpieza en seco, de productos textiles y de piel"),
    ("9602", "9602 - Peluquería y otros tratamientos de belleza"),
    ("9603", "9603 - Pompas fúnebres y actividades relacionadas"),
    ("9609", "9609 - Otras actividades de servicios personales n.c.p."),
    ("9700", "9700 - Actividades de los hogares individuales como empleadores de personal doméstico"),
    ("9810", "9810 - Actividades no diferenciadas de los hogares individuales como productores de bienes"),
    ("9820", "9820 - Actividades no diferenciadas de los hogares individuales como productores de servicios"),
    ("9900", "9900 - Actividades de organizaciones y entidades extraterritoriales"),
]

# Solo las etiquetas para el selectbox
CIUU_LABELS = [label for _, label in CIUU_OPCIONES]
CIUU_CODES  = [code  for code, _ in CIUU_OPCIONES]

# ==================== DOCUMENTOS REQUERIDOS ====================
DOCUMENTOS = {
    'doc_rut':          '1. RUT',
    'doc_ccio':         '2. C.CIO',
    'doc_rep_legal':    '3. C. Rep Legal',
    'doc_cert_banca':   '4. Cert. Bancaria',
    'doc_cert_comerc':  '5. Cert. Comercial',
    'doc_composicion':  '6. Composición Accionaria / Certificado',
    'doc_registro':     '7. Registro',
    'doc_trat_datos':   '8. Autori. Trat. Datos',
    'doc_aviso_priv':   '9. Aviso de Privacidad',
    'doc_basc':         '10. BASC o Equivalente',
    'doc_acuerdo_seg':  '10.1 Acuerdo Seguridad',
    'doc_codigo_etica': '11. Divulgación Código de Ética',
    'doc_risk':         '12. RISK / Compliance',
}
TOTAL_DOCS = len(DOCUMENTOS)


# ==================== BASE DE DATOS ====================
class DatabaseManager:
    def __init__(self):
        self.db_url = SUPABASE_DB_URL
        self.init_database()

    def get_connection(self):
        return psycopg2.connect(self.db_url)

    def init_database(self):
        try:
            conn = self.get_connection()
            cur = conn.cursor()

            cur.execute('''
                CREATE TABLE IF NOT EXISTS proveedores (
                    id SERIAL PRIMARY KEY,
                    fecha_registro TEXT,
                    nit_cedula TEXT,
                    nombre TEXT NOT NULL,
                    tipo_bien_servicio TEXT,
                    tipo_actividad TEXT,
                    direccion_ciudad TEXT,
                    telefono TEXT,
                    contacto TEXT,
                    correo TEXT,
                    doc_rut INTEGER DEFAULT 0,
                    doc_ccio INTEGER DEFAULT 0,
                    doc_rep_legal INTEGER DEFAULT 0,
                    doc_cert_banca INTEGER DEFAULT 0,
                    doc_cert_comerc INTEGER DEFAULT 0,
                    doc_composicion INTEGER DEFAULT 0,
                    doc_registro INTEGER DEFAULT 0,
                    doc_trat_datos INTEGER DEFAULT 0,
                    doc_aviso_priv INTEGER DEFAULT 0,
                    doc_basc INTEGER DEFAULT 0,
                    doc_acuerdo_seg INTEGER DEFAULT 0,
                    doc_codigo_etica INTEGER DEFAULT 0,
                    doc_risk INTEGER DEFAULT 0,
                    fecha_vinculacion TEXT,
                    ultima_actualizacion TEXT,
                    proxima_actualizacion TEXT,
                    eval_inicial_fecha TEXT,
                    eval_inicial_riesgo TEXT,
                    reevaluacion TEXT,
                    control_visitas TEXT,
                    envio_retroalimentacion TEXT,
                    otros_documentos TEXT
                )
            ''')

            cur.execute('''
                CREATE TABLE IF NOT EXISTS documentos_pdf (
                    id SERIAL PRIMARY KEY,
                    proveedor_id INTEGER NOT NULL REFERENCES proveedores(id) ON DELETE CASCADE,
                    doc_key TEXT NOT NULL,
                    filename TEXT NOT NULL,
                    filesize INTEGER,
                    contenido BYTEA NOT NULL,
                    subido_en TEXT NOT NULL
                )
            ''')

            for col_def in [
                "ADD COLUMN IF NOT EXISTS tipo_actividad TEXT",
                "ADD COLUMN IF NOT EXISTS fecha_vinculacion TEXT",
                "ADD COLUMN IF NOT EXISTS nit_cedula TEXT",
            ]:
                try:
                    cur.execute(f"ALTER TABLE proveedores {col_def}")
                except Exception:
                    pass

            conn.commit()
            conn.close()
        except Exception as e:
            st.error(f"Error inicializando base de datos: {e}")

    def guardar_proveedor(self, datos):
        try:
            conn = self.get_connection()
            cur = conn.cursor()
            hora_col = datetime.now() - timedelta(hours=5)
            fecha_actual = hora_col.strftime('%Y-%m-%d %H:%M:%S')
            cur.execute('''
                INSERT INTO proveedores (
                    fecha_registro, nit_cedula, nombre, tipo_bien_servicio, tipo_actividad,
                    direccion_ciudad, telefono, contacto, correo,
                    doc_rut, doc_ccio, doc_rep_legal, doc_cert_banca, doc_cert_comerc,
                    doc_composicion, doc_registro, doc_trat_datos, doc_aviso_priv,
                    doc_basc, doc_acuerdo_seg, doc_codigo_etica, doc_risk,
                    fecha_vinculacion, ultima_actualizacion, proxima_actualizacion,
                    eval_inicial_fecha, eval_inicial_riesgo,
                    reevaluacion, control_visitas, envio_retroalimentacion, otros_documentos
                ) VALUES (
                    %s,%s,%s,%s,%s,%s,%s,%s,%s,
                    %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                    %s,%s,%s,%s,%s,%s,%s,%s,%s
                ) RETURNING id
            ''', (
                fecha_actual,
                datos.get('nit_cedula', ''),
                datos['nombre'], datos['tipo_bien_servicio'], datos.get('tipo_actividad', ''),
                datos['direccion_ciudad'], datos['telefono'], datos['contacto'], datos['correo'],
                datos['doc_rut'], datos['doc_ccio'], datos['doc_rep_legal'],
                datos['doc_cert_banca'], datos['doc_cert_comerc'], datos['doc_composicion'],
                datos['doc_registro'], datos['doc_trat_datos'], datos['doc_aviso_priv'],
                datos['doc_basc'], datos['doc_acuerdo_seg'], datos['doc_codigo_etica'],
                datos['doc_risk'],
                datos.get('fecha_vinculacion', ''),
                datos['ultima_actualizacion'], datos['proxima_actualizacion'],
                datos['eval_inicial_fecha'], datos['eval_inicial_riesgo'],
                datos['reevaluacion'], datos['control_visitas'],
                datos['envio_retroalimentacion'], datos['otros_documentos'],
            ))
            result = cur.fetchone()
            conn.commit()
            conn.close()
            return result[0] if result else None
        except Exception as e:
            st.error(f"Error guardando proveedor: {e}")
            return None

    def actualizar_proveedor(self, proveedor_id, datos):
        try:
            conn = self.get_connection()
            cur = conn.cursor()
            cur.execute('''
                UPDATE proveedores SET
                    nit_cedula=%s, nombre=%s, tipo_bien_servicio=%s, tipo_actividad=%s,
                    direccion_ciudad=%s, telefono=%s, contacto=%s, correo=%s,
                    doc_rut=%s, doc_ccio=%s, doc_rep_legal=%s, doc_cert_banca=%s,
                    doc_cert_comerc=%s, doc_composicion=%s, doc_registro=%s,
                    doc_trat_datos=%s, doc_aviso_priv=%s, doc_basc=%s,
                    doc_acuerdo_seg=%s, doc_codigo_etica=%s, doc_risk=%s,
                    fecha_vinculacion=%s, ultima_actualizacion=%s, proxima_actualizacion=%s,
                    eval_inicial_fecha=%s, eval_inicial_riesgo=%s,
                    reevaluacion=%s, control_visitas=%s,
                    envio_retroalimentacion=%s, otros_documentos=%s
                WHERE id=%s
            ''', (
                datos.get('nit_cedula', ''),
                datos['nombre'], datos['tipo_bien_servicio'], datos.get('tipo_actividad', ''),
                datos['direccion_ciudad'], datos['telefono'], datos['contacto'], datos['correo'],
                datos['doc_rut'], datos['doc_ccio'], datos['doc_rep_legal'],
                datos['doc_cert_banca'], datos['doc_cert_comerc'], datos['doc_composicion'],
                datos['doc_registro'], datos['doc_trat_datos'], datos['doc_aviso_priv'],
                datos['doc_basc'], datos['doc_acuerdo_seg'], datos['doc_codigo_etica'],
                datos['doc_risk'],
                datos.get('fecha_vinculacion', ''),
                datos['ultima_actualizacion'], datos['proxima_actualizacion'],
                datos['eval_inicial_fecha'], datos['eval_inicial_riesgo'],
                datos['reevaluacion'], datos['control_visitas'],
                datos['envio_retroalimentacion'], datos['otros_documentos'],
                proveedor_id,
            ))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            st.error(f"Error actualizando: {e}")
            return False

    def obtener_proveedores(self):
        try:
            conn = self.get_connection()
            df = pd.read_sql_query("SELECT * FROM proveedores ORDER BY nombre", conn)
            conn.close()
            return df
        except Exception as e:
            st.error(f"Error obteniendo proveedores: {e}")
            return pd.DataFrame()

    def eliminar_proveedor(self, proveedor_id):
        conn = self.get_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM proveedores WHERE id = %s", (proveedor_id,))
        conn.commit()
        conn.close()

    def subir_pdf(self, proveedor_id: int, doc_key: str, filename: str, contenido: bytes):
        try:
            conn = self.get_connection()
            cur = conn.cursor()
            hora_col = datetime.now() - timedelta(hours=5)
            subido_en = hora_col.strftime('%Y-%m-%d %H:%M:%S')
            cur.execute('''
                INSERT INTO documentos_pdf (proveedor_id, doc_key, filename, filesize, contenido, subido_en)
                VALUES (%s, %s, %s, %s, %s, %s) RETURNING id
            ''', (proveedor_id, doc_key, filename, len(contenido),
                  psycopg2.Binary(contenido), subido_en))
            new_id = cur.fetchone()[0]
            conn.commit()
            conn.close()
            return new_id
        except Exception as e:
            st.error(f"Error subiendo PDF: {e}")
            return None

    def listar_versiones(self, proveedor_id: int, doc_key: str):
        try:
            conn = self.get_connection()
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute('''
                SELECT id, filename, filesize, subido_en
                FROM documentos_pdf
                WHERE proveedor_id = %s AND doc_key = %s
                ORDER BY subido_en DESC
            ''', (proveedor_id, doc_key))
            rows = [dict(r) for r in cur.fetchall()]
            conn.close()
            return rows
        except Exception:
            return []

    def descargar_pdf(self, pdf_id: int):
        try:
            conn = self.get_connection()
            cur = conn.cursor()
            cur.execute("SELECT filename, contenido FROM documentos_pdf WHERE id = %s", (pdf_id,))
            row = cur.fetchone()
            conn.close()
            if row:
                return row[0], bytes(row[1])
            return None, None
        except Exception as e:
            st.error(f"Error descargando: {e}")
            return None, None

    def eliminar_version_pdf(self, pdf_id: int):
        try:
            conn = self.get_connection()
            cur = conn.cursor()
            cur.execute("DELETE FROM documentos_pdf WHERE id = %s", (pdf_id,))
            conn.commit()
            conn.close()
        except Exception as e:
            st.error(f"Error eliminando versión: {e}")


# ==================== FUNCIONES AUXILIARES ====================
def calcular_indice(row):
    entregados = sum(1 for col in DOCUMENTOS if int(row.get(col) or 0) == 1)
    return round((entregados / TOTAL_DOCS) * 100, 1)

def color_indice(pct):
    return "🟢" if pct >= 80 else "🟡" if pct >= 50 else "🔴"

def estado_texto(pct):
    return "COMPLETO" if pct >= 80 else "EN PROCESO" if pct >= 50 else "CRÍTICO"

def fmt_bytes(size):
    if not size: return ""
    if size < 1024:        return f"{size} B"
    if size < 1_048_576:   return f"{size/1024:.1f} KB"
    return f"{size/1_048_576:.1f} MB"

def _parse_fecha(valor):
    if not valor or str(valor).strip() in ('', 'None', 'nan'):
        return None
    s = str(valor).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y'):
        try:
            return datetime.strptime(s[:10], fmt[:10])
        except ValueError:
            continue
    return None

def _dias_diferencia(fecha_str):
    dt = _parse_fecha(fecha_str)
    if dt is None:
        return None
    return (dt - datetime.now()).days

def get_ciuu_index(valor_actual):
    """Retorna el índice en CIUU_LABELS para el valor guardado."""
    if not valor_actual or str(valor_actual).strip() in ('', 'None', 'nan'):
        return 0
    val = str(valor_actual).strip()
    # Buscar por etiqueta exacta
    for i, label in enumerate(CIUU_LABELS):
        if label == val:
            return i
    # Buscar si el valor guardado empieza con el código
    for i, (code, label) in enumerate(CIUU_OPCIONES):
        if val.startswith(code) or label.startswith(val):
            return i
    return 0


# ==================== WIDGET PDF POR DOCUMENTO ====================
def widget_documento_pdf(db: DatabaseManager, proveedor_id: int,
                          doc_key: str, doc_label: str,
                          checked: bool, form_key_prefix: str) -> int:
    col_chk, col_body = st.columns([1, 9])
    with col_chk:
        entregado = st.checkbox(
            "✔", value=checked,
            key=f"{form_key_prefix}_{doc_key}_chk",
            label_visibility="collapsed",
        )
    with col_body:
        icon = "✅" if entregado else "📄"
        with st.expander(f"{icon}  {doc_label}", expanded=False):

            st.markdown("##### 📤 Subir nueva versión")
            uploaded = st.file_uploader(
                "Selecciona PDF", type=["pdf"],
                key=f"{form_key_prefix}_{doc_key}_uploader",
                label_visibility="collapsed",
            )
            if uploaded:
                if st.button("💾 Guardar esta versión",
                             key=f"{form_key_prefix}_{doc_key}_btn"):
                    contenido = uploaded.read()
                    new_id = db.subir_pdf(proveedor_id, doc_key, uploaded.name, contenido)
                    if new_id:
                        st.success(f"✅ **{uploaded.name}** guardado ({fmt_bytes(len(contenido))})")
                        st.rerun()

            versiones = db.listar_versiones(proveedor_id, doc_key)
            if versiones:
                st.markdown(f"##### 📂 Historial — {len(versiones)} versión(es)")
                for v in versiones:
                    vc1, vc2, vc3, vc4 = st.columns([4, 2, 2, 1])
                    with vc1:
                        st.markdown(f"🗂 `{v['filename']}`")
                        st.caption(v['subido_en'])
                    with vc2:
                        st.caption(fmt_bytes(v['filesize']))
                    with vc3:
                        fname, fbytes = db.descargar_pdf(v['id'])
                        if fbytes:
                            st.download_button(
                                label="⬇️ Descargar", data=fbytes,
                                file_name=fname, mime="application/pdf",
                                key=f"dl_{v['id']}",
                            )
                    with vc4:
                        if st.button("🗑️", key=f"delpdf_{v['id']}", help="Eliminar esta versión"):
                            db.eliminar_version_pdf(v['id'])
                            st.rerun()
            else:
                st.caption("📭 Sin archivos subidos aún")

    return 1 if entregado else 0


# ==================== GENERADOR EXCEL ====================
def generar_excel_proveedores(df):
    output = io.BytesIO()
    wb = Workbook()

    # ── Estilos compartidos ──────────────────────────────────────────────────
    h_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    h_font = Font(color="FFFFFF", bold=True, size=11)
    verde  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    rojo   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    amari  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    azul_c = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    gris_c = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    borde  = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))
    centro = Alignment(horizontal='center', vertical='center')
    wrap_c = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def hdr(ws, row, col, value):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = h_font; cell.fill = h_fill
        cell.alignment = wrap_c; cell.border = borde
        return cell

    # ════════════════════════════════════════════════════════════════════════
    # Hoja 1 – Directorio
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Directorio Proveedores"
    ws1.merge_cells('A1:I1')
    ws1['A1'] = "DIRECTORIO DE PROVEEDORES"
    ws1['A1'].font = Font(size=15, bold=True, color="1F4E78")
    ws1['A1'].alignment = centro
    ws1.row_dimensions[1].height = 30
    ws1.merge_cells('A2:I2')
    ws1['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   Total: {len(df)}"
    ws1['A2'].alignment = centro
    ws1['A2'].font = Font(italic=True, color="555555")
    for c, h in enumerate(['NIT / Cédula', 'Nombre Proveedor', 'Tipo Bien / Servicio', 'Tipo Actividad (CIUU)',
                            'Dirección / Ciudad', 'Teléfono', 'Contacto', 'Correo', 'Fecha Registro'], 1):
        hdr(ws1, 4, c, h)
    ws1.row_dimensions[4].height = 20
    for r, (_, row) in enumerate(df.iterrows(), 5):
        for c, f in enumerate(['nit_cedula', 'nombre', 'tipo_bien_servicio', 'tipo_actividad',
                                'direccion_ciudad', 'telefono', 'contacto', 'correo', 'fecha_registro'], 1):
            cell = ws1.cell(row=r, column=c, value=str(row.get(f, '') or ''))
            cell.border = borde
            cell.fill = azul_c if r % 2 == 0 else PatternFill()
    for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'], [18, 35, 25, 30, 28, 18, 25, 30, 20]):
        ws1.column_dimensions[col].width = w

    # ════════════════════════════════════════════════════════════════════════
    # Hoja 2 – Documentos y Cumplimiento
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Documentos y Cumplimiento")
    total_cols = 2 + TOTAL_DOCS + 3
    ws2.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    ws2['A1'] = "ESTADO DOCUMENTAL POR PROVEEDOR"
    ws2['A1'].font = Font(size=15, bold=True, color="1F4E78")
    ws2['A1'].alignment = centro
    hdrs2 = (['Proveedor', '% Índice'] + list(DOCUMENTOS.values()) +
             ['Fecha Vinculación', 'Última Actualización', 'Próxima Actualización'])
    for c, h in enumerate(hdrs2, 1):
        hdr(ws2, 3, c, h)
    ws2.row_dimensions[3].height = 55
    for r, (_, row) in enumerate(df.iterrows(), 4):
        ind = calcular_indice(row)
        c1 = ws2.cell(row=r, column=1, value=str(row.get('nombre', '')))
        c1.border = borde; c1.font = Font(bold=True)
        c2 = ws2.cell(row=r, column=2, value=f"{ind}%")
        c2.alignment = centro; c2.border = borde; c2.font = Font(bold=True)
        c2.fill = verde if ind >= 80 else amari if ind >= 50 else rojo
        for ci, key in enumerate(DOCUMENTOS.keys(), 3):
            val = int(row.get(key) or 0)
            cell = ws2.cell(row=r, column=ci, value="✓ SÍ" if val else "✗ NO")
            cell.alignment = centro; cell.border = borde
            cell.fill = verde if val else rojo
        col_ua = 3 + TOTAL_DOCS
        ws2.cell(row=r, column=col_ua,   value=str(row.get('fecha_vinculacion', '') or '')).border = borde
        ws2.cell(row=r, column=col_ua+1, value=str(row.get('ultima_actualizacion', '') or '')).border = borde
        ws2.cell(row=r, column=col_ua+2, value=str(row.get('proxima_actualizacion', '') or '')).border = borde
    ws2.column_dimensions['A'].width = 32
    ws2.column_dimensions['B'].width = 14
    for i in range(3, total_cols + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 13

    # ════════════════════════════════════════════════════════════════════════
    # Hoja 3 – Evaluaciones y Control
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Evaluaciones y Control")
    ws3.merge_cells('A1:G1')
    ws3['A1'] = "EVALUACIONES Y CONTROL DE PROVEEDORES"
    ws3['A1'].font = Font(size=15, bold=True, color="1F4E78")
    ws3['A1'].alignment = centro
    ws3.row_dimensions[1].height = 30
    for c, h in enumerate(['Proveedor', '13. Nivel de Riesgo', '13. Fecha Eval.',
                            '14. Reevaluación', '15. Control Visitas',
                            '16. Retroalimentación', '17. Otros Docs'], 1):
        hdr(ws3, 3, c, h)
    ws3.row_dimensions[3].height = 40
    riesgo_color = {'ALTO': rojo, 'MEDIO': amari, 'BAJO': verde}
    for r, (_, row) in enumerate(df.iterrows(), 4):
        riesgo = str(row.get('eval_inicial_riesgo', '') or '')
        for ci, v in enumerate([row.get('nombre', ''), riesgo,
                                  row.get('eval_inicial_fecha', ''), row.get('reevaluacion', ''),
                                  row.get('control_visitas', ''), row.get('envio_retroalimentacion', ''),
                                  row.get('otros_documentos', '')], 1):
            cell = ws3.cell(row=r, column=ci, value=str(v) if v else '')
            cell.border = borde
            if ci == 2 and riesgo in riesgo_color:
                cell.fill = riesgo_color[riesgo]
                cell.font = Font(bold=True)
                cell.alignment = centro
    for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G'], [32, 20, 18, 24, 24, 24, 28]):
        ws3.column_dimensions[col].width = w

    # ════════════════════════════════════════════════════════════════════════
    # Hoja 4 – Informe Ejecutivo
    # ════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Informe Ejecutivo")
    ws4.merge_cells('A1:F1')
    ws4['A1'] = "INFORME EJECUTIVO — GESTIÓN DE PROVEEDORES"
    ws4['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws4['A1'].fill = h_fill; ws4['A1'].alignment = centro
    ws4.row_dimensions[1].height = 35
    ws4.merge_cells('A2:F2')
    ws4['A2'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   Total: {len(df)}"
    ws4['A2'].alignment = centro; ws4['A2'].font = Font(italic=True, color="444444")

    indices_list  = [calcular_indice(row) for _, row in df.iterrows()]
    prom_indice   = sum(indices_list) / len(indices_list) if indices_list else 0
    n_criticos    = sum(1 for i in indices_list if i < 50)
    n_proceso     = sum(1 for i in indices_list if 50 <= i < 80)
    n_completos   = sum(1 for i in indices_list if i >= 80)
    pct_completos = round(n_completos / len(df) * 100, 1) if len(df) > 0 else 0

    ws4.cell(row=4, column=1).value = "INDICADORES CLAVE DE DESEMPEÑO"
    ws4.cell(row=4, column=1).font = Font(bold=True, size=12, color="1F4E78")
    ws4.merge_cells('A4:F4'); ws4.row_dimensions[4].height = 22

    for i, (label, valor, nivel) in enumerate([
        ("Total Proveedores Registrados",            len(df),               None),
        ("Índice Promedio de Cumplimiento",          f"{prom_indice:.1f}%", prom_indice),
        ("Proveedores Completos (≥ 80%)",            n_completos,           100),
        ("Proveedores En Proceso (50–79%)",          n_proceso,             50),
        ("Proveedores Críticos (< 50%)",             n_criticos,            0),
        ("% Proveedores Completamente Certificados", f"{pct_completos:.1f}%", pct_completos),
    ], 5):
        cl = ws4.cell(row=i, column=1, value=label)
        cl.font = Font(bold=True); cl.border = borde
        cl.fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
        ws4.merge_cells(f'A{i}:D{i}')
        cv = ws4.cell(row=i, column=5, value=valor)
        cv.alignment = centro; cv.border = borde; cv.font = Font(bold=True, size=12)
        if nivel is not None:
            cv.fill = verde if nivel >= 80 else amari if nivel >= 50 else rojo
        ws4.merge_cells(f'E{i}:F{i}')

    row_rank = 12
    ws4.cell(row=row_rank, column=1).value = "RANKING DE CUMPLIMIENTO"
    ws4.cell(row=row_rank, column=1).font = Font(bold=True, size=12, color="1F4E78")
    ws4.merge_cells(f'A{row_rank}:F{row_rank}')
    for c, h in enumerate(['#', 'Proveedor', 'Tipo Bien', '% Cumplimiento', 'Estado', 'Docs'], 1):
        hdr(ws4, row_rank + 1, c, h)

    df_rk = df.copy()
    df_rk['_idx'] = indices_list
    df_rk = df_rk.sort_values('_idx', ascending=False).reset_index(drop=True)
    for ri, (_, row) in enumerate(df_rk.iterrows(), row_rank + 2):
        ind = row['_idx']
        docs_ok = sum(1 for k in DOCUMENTOS if int(row.get(k) or 0) == 1)
        estado = "✅ COMPLETO" if ind >= 80 else "⚠️ EN PROCESO" if ind >= 50 else "❌ CRÍTICO"
        for ci, v in enumerate([ri - row_rank - 1, row.get('nombre', ''),
                                  row.get('tipo_bien_servicio', ''),
                                  f"{ind}%", estado, f"{docs_ok}/{TOTAL_DOCS}"], 1):
            cell = ws4.cell(row=ri, column=ci, value=v)
            cell.border = borde
            if ci == 4:
                cell.alignment = centro; cell.font = Font(bold=True)
                cell.fill = verde if ind >= 80 else amari if ind >= 50 else rojo
            elif ci == 5:
                cell.alignment = centro

    row_doc = row_rank + len(df_rk) + 4
    ws4.cell(row=row_doc, column=1).value = "ANÁLISIS DE ENTREGA POR DOCUMENTO"
    ws4.cell(row=row_doc, column=1).font = Font(bold=True, size=12, color="1F4E78")
    ws4.merge_cells(f'A{row_doc}:F{row_doc}')
    for c, h in enumerate(['Documento', 'Con Doc.', 'Total', '% Entrega', 'Faltantes'], 1):
        hdr(ws4, row_doc + 1, c, h)
    for di, (key, label) in enumerate(DOCUMENTOS.items(), row_doc + 2):
        entregados = int(df[key].sum()) if key in df.columns else 0
        faltantes  = len(df) - entregados
        pct_doc    = round(entregados / len(df) * 100, 1) if len(df) > 0 else 0
        for ci, v in enumerate([label, entregados, len(df), f"{pct_doc}%", faltantes], 1):
            cell = ws4.cell(row=di, column=ci, value=v)
            cell.border = borde
            if ci == 4:
                cell.alignment = centro; cell.font = Font(bold=True)
                cell.fill = verde if pct_doc >= 80 else amari if pct_doc >= 50 else rojo
    for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F'], [38, 22, 18, 16, 18, 18]):
        ws4.column_dimensions[col].width = w

    # ════════════════════════════════════════════════════════════════════════
    # Hoja 5 – Trazabilidad de Actualizaciones
    # ════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Trazabilidad Actualizaciones")
    total_prov = len(df)

    ws5.merge_cells('A1:H1')
    ws5['A1'] = "TRAZABILIDAD DE ACTUALIZACIONES"
    ws5['A1'].font = Font(size=15, bold=True, color="FFFFFF")
    ws5['A1'].fill = h_fill
    ws5['A1'].alignment = centro
    ws5.row_dimensions[1].height = 32

    ws5.merge_cells('A2:H2')
    ws5['A2'] = (f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   "
                 f"Total proveedores: {total_prov}")
    ws5['A2'].alignment = centro
    ws5['A2'].font = Font(italic=True, color="444444")

    ws5.merge_cells('A4:H4')
    ws5['A4'] = "▌ RESUMEN EJECUTIVO DE ACTUALIZACIONES"
    ws5['A4'].font = Font(bold=True, size=12, color="1F4E78")
    ws5['A4'].fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    ws5.row_dimensions[4].height = 22

    n_al_dia, n_por_vencer, n_vencidos, n_sin_fecha = 0, 0, 0, 0
    for _, row in df.iterrows():
        dias = _dias_diferencia(row.get('proxima_actualizacion', ''))
        if dias is None:          n_sin_fecha  += 1
        elif dias > 30:           n_al_dia     += 1
        elif dias >= 0:           n_por_vencer += 1
        else:                     n_vencidos   += 1

    pct_al_dia     = round(n_al_dia     / total_prov * 100, 1) if total_prov else 0
    pct_por_vencer = round(n_por_vencer / total_prov * 100, 1) if total_prov else 0
    pct_vencidos   = round(n_vencidos   / total_prov * 100, 1) if total_prov else 0
    pct_sin_fecha  = round(n_sin_fecha  / total_prov * 100, 1) if total_prov else 0

    for c, h in enumerate(['Estado', 'N° Proveedores', '% del Total', 'Barra Visual'], 1):
        hdr(ws5, 5, c, h)

    resumen_rows = [
        ("🟢 Al día (próxima actualización > 30 días)",     n_al_dia,     pct_al_dia,     verde),
        ("🟡 Por vencer (próxima actualización ≤ 30 días)", n_por_vencer, pct_por_vencer, amari),
        ("🔴 Vencidos (fecha ya pasó)",                     n_vencidos,   pct_vencidos,   rojo),
        ("⚪ Sin fecha registrada",                          n_sin_fecha,  pct_sin_fecha,  gris_c),
    ]
    for ri, (estado, cantidad, pct, fill) in enumerate(resumen_rows, 6):
        barra = "■" * int(pct / 5) + "□" * (20 - int(pct / 5))
        for ci, v in enumerate([estado, cantidad, f"{pct}%", barra], 1):
            cell = ws5.cell(row=ri, column=ci, value=v)
            cell.border = borde; cell.fill = fill
            if ci in (2, 3):
                cell.alignment = centro; cell.font = Font(bold=True)

    for ci, v in enumerate(["TOTAL", total_prov, "100%", ""], 1):
        cell = ws5.cell(row=10, column=ci, value=v)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = h_fill; cell.alignment = centro; cell.border = borde

    ws5.merge_cells('A12:H12')
    ws5['A12'] = "▌ DETALLE POR PROVEEDOR"
    ws5['A12'].font = Font(bold=True, size=12, color="1F4E78")
    ws5['A12'].fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    ws5.row_dimensions[12].height = 22

    det_hdrs = ['Proveedor', 'Última Actualización', 'Próxima Actualización',
                'Días para Vencer', 'Estado Vigencia', '% Docs Entregados',
                'N° Actualizaciones', 'Observación']
    for c, h in enumerate(det_hdrs, 1):
        hdr(ws5, 13, c, h)
    ws5.row_dimensions[13].height = 45

    for r, (_, row) in enumerate(df.iterrows(), 14):
        nombre   = str(row.get('nombre', '') or '')
        ult_act  = str(row.get('ultima_actualizacion', '') or '').strip()
        prox_act = str(row.get('proxima_actualizacion', '') or '').strip()
        indice   = calcular_indice(row)
        dias     = _dias_diferencia(prox_act)
        tiene_ult = ult_act not in ('', 'None', 'nan')
        n_act     = 1 if tiene_ult else 0

        if dias is None:
            estado_vig = "⚪ Sin fecha";          fill_est = gris_c
        elif dias > 30:
            estado_vig = "🟢 Al día";             fill_est = verde
        elif dias >= 0:
            estado_vig = f"🟡 Vence en {dias}d";  fill_est = amari
        else:
            estado_vig = f"🔴 Vencido ({abs(dias)}d)"; fill_est = rojo

        if dias is None:
            obs = "Sin fechas — requiere ingreso"
        elif dias < 0:
            obs = f"Vencido hace {abs(dias)} días — actualización urgente"
        elif dias <= 30:
            obs = f"Vence en {dias} días — programar actualización"
        else:
            obs = "Vigente"

        vals = [
            nombre,
            ult_act  if tiene_ult                           else "—",
            prox_act if prox_act not in ('', 'None', 'nan') else "—",
            dias     if dias is not None                    else "—",
            estado_vig,
            f"{indice}%",
            n_act,
            obs,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws5.cell(row=r, column=c, value=v)
            cell.border = borde
            if c == 4 and isinstance(v, int):
                cell.alignment = centro; cell.font = Font(bold=True)
                cell.fill = verde if v > 30 else amari if v >= 0 else rojo
            elif c == 5:
                cell.fill = fill_est; cell.alignment = centro; cell.font = Font(bold=True)
            elif c == 6:
                cell.alignment = centro; cell.font = Font(bold=True)
                cell.fill = verde if indice >= 80 else amari if indice >= 50 else rojo
            elif c == 7:
                cell.alignment = centro; cell.font = Font(bold=True)
                cell.fill = verde if n_act >= 1 else rojo
            elif r % 2 == 0 and c not in (4, 5, 6, 7):
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    sep = 13 + len(df) + 2
    ws5.merge_cells(f'A{sep}:H{sep}')
    ws5[f'A{sep}'] = "▌ ESTADÍSTICAS GENERALES"
    ws5[f'A{sep}'].font = Font(bold=True, size=12, color="1F4E78")
    ws5[f'A{sep}'].fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    ws5.row_dimensions[sep].height = 22

    tiene_ultima  = sum(1 for _, r in df.iterrows()
                        if str(r.get('ultima_actualizacion', '') or '').strip()
                        not in ('', 'None', 'nan'))
    tiene_proxima = sum(1 for _, r in df.iterrows()
                        if str(r.get('proxima_actualizacion', '') or '').strip()
                        not in ('', 'None', 'nan'))
    pct_ult  = round(tiene_ultima  / total_prov * 100, 1) if total_prov else 0
    pct_prox = round(tiene_proxima / total_prov * 100, 1) if total_prov else 0

    for c, h in enumerate(['Indicador', 'Cantidad', '% del Total', 'Observación'], 1):
        hdr(ws5, sep + 1, c, h)

    stats = [
        ("Proveedores con Última Actualización registrada",
         tiene_ultima, pct_ult,
         "Registro histórico presente" if pct_ult == 100 else "Faltan registros"),
        ("Proveedores con Próxima Actualización programada",
         tiene_proxima, pct_prox,
         "Seguimiento programado" if pct_prox == 100 else "Sin programar"),
        ("Proveedores actualizados al menos 1 vez",
         tiene_ultima, pct_ult,
         "≥ 1 actualización documentada"),
        ("Total actualizaciones registradas en el sistema",
         tiene_ultima, "—",
         "Suma de todas las actualizaciones únicas"),
    ]
    for ri, (ind, cant, pct, obs) in enumerate(stats, sep + 2):
        for ci, v in enumerate([ind, cant, f"{pct}%" if isinstance(pct, float) else pct, obs], 1):
            cell = ws5.cell(row=ri, column=ci, value=v)
            cell.border = borde
            if ci == 3 and isinstance(pct, float):
                cell.alignment = centro; cell.font = Font(bold=True)
                cell.fill = verde if pct >= 80 else amari if pct >= 50 else rojo
            if ri % 2 == 0 and ci != 3:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'],
                      [38, 22, 22, 18, 22, 20, 20, 42]):
        ws5.column_dimensions[col].width = w

    # ════════════════════════════════════════════════════════════════════════
    # Hoja 6 – Análisis por Tipo de Actividad (CIUU)  ← NUEVA
    # ════════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Análisis por Actividad CIUU")

    ws6.merge_cells('A1:G1')
    ws6['A1'] = "ANÁLISIS DE CUMPLIMIENTO POR TIPO DE ACTIVIDAD (CIUU)"
    ws6['A1'].font = Font(size=15, bold=True, color="FFFFFF")
    ws6['A1'].fill = h_fill
    ws6['A1'].alignment = centro
    ws6.row_dimensions[1].height = 32

    ws6.merge_cells('A2:G2')
    ws6['A2'] = (f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   "
                 f"Total proveedores: {len(df)}")
    ws6['A2'].alignment = centro
    ws6['A2'].font = Font(italic=True, color="444444")

    # ── Bloque A: Resumen por tipo de actividad ─────────────────────────────
    ws6.merge_cells('A4:G4')
    ws6['A4'] = "▌ PORCENTAJE DE CUMPLIMIENTO DOCUMENTAL POR TIPO DE ACTIVIDAD"
    ws6['A4'].font = Font(bold=True, size=12, color="1F4E78")
    ws6['A4'].fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    ws6.row_dimensions[4].height = 22

    for c, h in enumerate(['Tipo de Actividad (CIUU)', 'N° Proveedores',
                            '% Prom. Cumplimiento', 'Completos (≥80%)',
                            'En Proceso (50-79%)', 'Críticos (<50%)', 'Estado General'], 1):
        hdr(ws6, 5, c, h)
    ws6.row_dimensions[5].height = 40

    # Agrupar proveedores por tipo_actividad
    df_act = df.copy()
    df_act['_indice'] = [calcular_indice(r) for _, r in df_act.iterrows()]
    df_act['_actividad'] = df_act['tipo_actividad'].fillna('').replace('', 'Sin Actividad Registrada')

    grupos = df_act.groupby('_actividad')

    actividades_data = []
    for actividad, grupo in grupos:
        n_prov       = len(grupo)
        indices_g    = grupo['_indice'].tolist()
        prom_g       = round(sum(indices_g) / n_prov, 1) if n_prov > 0 else 0
        n_comp_g     = sum(1 for i in indices_g if i >= 80)
        n_proc_g     = sum(1 for i in indices_g if 50 <= i < 80)
        n_crit_g     = sum(1 for i in indices_g if i < 50)
        actividades_data.append((actividad, n_prov, prom_g, n_comp_g, n_proc_g, n_crit_g))

    # Ordenar por promedio descendente
    actividades_data.sort(key=lambda x: x[2], reverse=True)

    for ri, (act, n_prov, prom_g, n_comp_g, n_proc_g, n_crit_g) in enumerate(actividades_data, 6):
        estado_g = "✅ BUENO" if prom_g >= 80 else "⚠️ REGULAR" if prom_g >= 50 else "❌ CRÍTICO"
        fill_g   = verde if prom_g >= 80 else amari if prom_g >= 50 else rojo
        vals = [act, n_prov, f"{prom_g}%", n_comp_g, n_proc_g, n_crit_g, estado_g]
        for ci, v in enumerate(vals, 1):
            cell = ws6.cell(row=ri, column=ci, value=v)
            cell.border = borde
            if ci == 3:
                cell.alignment = centro; cell.font = Font(bold=True)
                cell.fill = fill_g
            elif ci == 7:
                cell.alignment = centro; cell.font = Font(bold=True)
                cell.fill = fill_g
            elif ci in (2, 4, 5, 6):
                cell.alignment = centro
            if ri % 2 == 0 and ci not in (3, 7):
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Fila total
    row_tot = 6 + len(actividades_data)
    for ci, v in enumerate(["TOTAL GENERAL", len(df),
                             f"{round(sum(i[2]*i[1] for i in actividades_data)/len(df),1) if len(df)>0 else 0}%",
                             sum(i[3] for i in actividades_data),
                             sum(i[4] for i in actividades_data),
                             sum(i[5] for i in actividades_data), ""], 1):
        cell = ws6.cell(row=row_tot, column=ci, value=v)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = h_fill; cell.alignment = centro; cell.border = borde

    # ── Bloque B: Detalle individual por proveedor agrupado por actividad ───
    row_det = row_tot + 2
    ws6.merge_cells(f'A{row_det}:G{row_det}')
    ws6[f'A{row_det}'] = "▌ DETALLE INDIVIDUAL POR PROVEEDOR Y ACTIVIDAD"
    ws6[f'A{row_det}'].font = Font(bold=True, size=12, color="1F4E78")
    ws6[f'A{row_det}'].fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    ws6.row_dimensions[row_det].height = 22

    for c, h in enumerate(['Tipo de Actividad (CIUU)', 'NIT / Cédula', 'Nombre Proveedor',
                            '% Cumplimiento', 'Docs Entregados', 'Estado', 'Nivel de Riesgo'], 1):
        hdr(ws6, row_det + 1, c, h)

    row_cur = row_det + 2
    for act, _, _, _, _, _ in actividades_data:
        grupo = df_act[df_act['_actividad'] == act]
        grupo_sorted = grupo.sort_values('_indice', ascending=False)
        for _, prow in grupo_sorted.iterrows():
            ind_p    = prow['_indice']
            docs_ok  = sum(1 for k in DOCUMENTOS if int(prow.get(k) or 0) == 1)
            estado_p = "✅ COMPLETO" if ind_p >= 80 else "⚠️ EN PROCESO" if ind_p >= 50 else "❌ CRÍTICO"
            fill_p   = verde if ind_p >= 80 else amari if ind_p >= 50 else rojo
            riesgo_p = str(prow.get('eval_inicial_riesgo', '') or '')
            fill_r   = rojo if riesgo_p == 'ALTO' else amari if riesgo_p == 'MEDIO' else verde if riesgo_p == 'BAJO' else gris_c

            vals = [act, str(prow.get('nit_cedula', '') or ''), str(prow.get('nombre', '') or ''),
                    f"{ind_p}%", f"{docs_ok}/{TOTAL_DOCS}", estado_p, riesgo_p if riesgo_p else '—']
            for ci, v in enumerate(vals, 1):
                cell = ws6.cell(row=row_cur, column=ci, value=v)
                cell.border = borde
                if ci == 4:
                    cell.alignment = centro; cell.font = Font(bold=True); cell.fill = fill_p
                elif ci == 5:
                    cell.alignment = centro
                elif ci == 6:
                    cell.alignment = centro; cell.font = Font(bold=True); cell.fill = fill_p
                elif ci == 7:
                    cell.alignment = centro; cell.font = Font(bold=True); cell.fill = fill_r
                elif row_cur % 2 == 0 and ci not in (4, 6, 7):
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            row_cur += 1

    for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G'],
                      [40, 18, 35, 18, 16, 16, 16]):
        ws6.column_dimensions[col].width = w

    wb.save(output)
    output.seek(0)
    return output


# ==================== MAIN ====================
def main():
    st.set_page_config(page_title="Gestión de Proveedores", layout="wide", page_icon="🏢")
    st.title("🏢 Sistema de Gestión de Proveedores")
    st.markdown("**Control Documental, Evaluación y Trazabilidad de Proveedores**")

    if 'db' not in st.session_state:
        with st.spinner("Conectando a la base de datos..."):
            st.session_state.db = DatabaseManager()
    db = st.session_state.db

    tab1, tab2, tab3 = st.tabs([
        "➕ Nuevo Proveedor",
        "📋 Lista de Proveedores",
        "📊 Reportes y Exportación",
    ])

    # ===========================================================
    # TAB 1 – NUEVO PROVEEDOR
    # ===========================================================
    with tab1:
        st.header("Registro de Nuevo Proveedor")

        with st.form("form_proveedor", clear_on_submit=True):
            st.subheader("📌 Información General")
            col1, col2 = st.columns(2)
            with col1:
                nit_cedula         = st.text_input("NIT / Cédula *", placeholder="Ej: 900123456-1")
                nombre             = st.text_input("Nombre del Proveedor *", placeholder="Razón social")
                tipo_bien_servicio = st.text_input("Tipo de Bien / Servicio *",
                                                    placeholder="Ej: Repuestos, Transporte…")
                tipo_actividad     = st.selectbox(
                    "Tipo de Actividad (CIUU)",
                    options=CIUU_LABELS,
                    index=0,
                )
            with col2:
                direccion_ciudad = st.text_input("Dirección / Ciudad",
                                                  placeholder="Ej: Cra 7 # 10-20, Bogotá")
                telefono = st.text_input("Teléfono / Celular")
                contacto = st.text_input("Contacto")
                correo   = st.text_input("Correo Electrónico")

            st.divider()
            st.subheader("📄 Documentos Solicitados")
            st.info(
                "💡 Marca los documentos recibidos. "
                "Para subir los PDFs, guarda primero el proveedor y luego edítalo desde la lista.",
                icon="ℹ️",
            )
            doc_values = {}
            cols_doc = st.columns(3)
            for idx, (key, label) in enumerate(DOCUMENTOS.items()):
                with cols_doc[idx % 3]:
                    doc_values[key] = 1 if st.checkbox(label, key=f"new_{key}") else 0

            docs_marcados  = sum(doc_values.values())
            indice_preview = round((docs_marcados / TOTAL_DOCS) * 100, 1)
            cp1, cp2 = st.columns(2)
            with cp1:
                st.metric("📊 Índice Documental", f"{indice_preview}%",
                          help=f"{docs_marcados} de {TOTAL_DOCS} marcados")
            with cp2:
                if   indice_preview >= 80: st.success("🟢 COMPLETO")
                elif indice_preview >= 50: st.warning("🟡 EN PROCESO")
                else:                      st.error("🔴 CRÍTICO")

            st.divider()
            st.subheader("📅 Fechas")
            cf1, cf2, cf3 = st.columns(3)
            with cf1: fecha_vinculacion    = st.date_input("📎 Fecha de Vinculación",   value=None)
            with cf2: ultima_actualizacion  = st.date_input("🔄 Última Actualización",   value=None)
            with cf3: proxima_actualizacion = st.date_input("⏭️ Próxima Actualización",  value=None)

            st.divider()
            st.subheader("🔍 Evaluaciones y Control")
            ce1, ce2 = st.columns(2)
            with ce1:
                eval_inicial_fecha  = st.date_input("13. Evaluación Inicial — Fecha", value=None)
                eval_inicial_riesgo = st.selectbox("13. Nivel de Riesgo del Proveedor",
                                                    ["", "BAJO", "MEDIO", "ALTO"])
                reevaluacion        = st.text_input("14. Reevaluación")
            with ce2:
                control_visitas         = st.text_input("15. Control de Visitas")
                envio_retroalimentacion = st.text_input("16. Envío Retroalimentación")
                otros_documentos        = st.text_area("17. Otros Documentos", height=80)

            st.divider()
            submit = st.form_submit_button("💾 Guardar Proveedor", type="primary")

            if submit:
                if not nombre.strip():
                    st.error("⚠️ El nombre del proveedor es obligatorio.")
                else:
                    datos = {
                        'nit_cedula': nit_cedula.strip(),
                        'nombre': nombre.strip().upper(),
                        'tipo_bien_servicio': tipo_bien_servicio,
                        'tipo_actividad': tipo_actividad if tipo_actividad != CIUU_LABELS[0] else '',
                        'direccion_ciudad': direccion_ciudad,
                        'telefono': telefono, 'contacto': contacto, 'correo': correo,
                        **doc_values,
                        'fecha_vinculacion':     str(fecha_vinculacion)     if fecha_vinculacion    else '',
                        'ultima_actualizacion':  str(ultima_actualizacion)  if ultima_actualizacion  else '',
                        'proxima_actualizacion': str(proxima_actualizacion) if proxima_actualizacion else '',
                        'eval_inicial_fecha':    str(eval_inicial_fecha)    if eval_inicial_fecha    else '',
                        'eval_inicial_riesgo': eval_inicial_riesgo,
                        'reevaluacion': reevaluacion,
                        'control_visitas': control_visitas,
                        'envio_retroalimentacion': envio_retroalimentacion,
                        'otros_documentos': otros_documentos,
                    }
                    prov_id = db.guardar_proveedor(datos)
                    if prov_id:
                        st.success(
                            f"✅ **{datos['nombre']}** guardado (ID: {prov_id})  |  "
                            f"Índice: **{indice_preview}%**"
                        )
                        st.info("📤 Ve a **Lista de Proveedores → Editar** para subir los PDFs.")
                        st.balloons()

    # ===========================================================
    # TAB 2 – LISTA DE PROVEEDORES
    # ===========================================================
    with tab2:
        st.header("📋 Lista de Proveedores")
        if st.button("🔄 Actualizar lista"):
            st.rerun()

        df = db.obtener_proveedores()

        if df.empty:
            st.info("No hay proveedores registrados aún.")
        else:
            indices_all = [calcular_indice(r) for _, r in df.iterrows()]
            prom_all    = sum(indices_all) / len(indices_all) if indices_all else 0

            c1, c2, c3, c4, c5 = st.columns(5)
            with c1: st.metric("Total",          len(df))
            with c2: st.metric("Índice Promedio", f"{prom_all:.1f}%")
            with c3: st.metric("🔴 Críticos",     sum(1 for i in indices_all if i < 50))
            with c4: st.metric("🟡 En Proceso",   sum(1 for i in indices_all if 50 <= i < 80))
            with c5: st.metric("🟢 Completos",    sum(1 for i in indices_all if i >= 80))

            st.divider()

            df_show = df[['id', 'nit_cedula', 'nombre', 'tipo_bien_servicio', 'tipo_actividad',
                           'direccion_ciudad', 'telefono', 'contacto', 'correo']].copy()
            df_show['% Docs'] = [f"{calcular_indice(r):.1f}%" for _, r in df.iterrows()]
            df_show['Estado'] = [
                f"{color_indice(calcular_indice(r))} {estado_texto(calcular_indice(r))}"
                for _, r in df.iterrows()
            ]
            df_show.columns = ['ID', 'NIT/Cédula', 'Nombre', 'Tipo', 'Actividad', 'Dirección',
                                'Teléfono', 'Contacto', 'Correo', '% Docs', 'Estado']
            st.dataframe(df_show, use_container_width=True, hide_index=True, height=360)

            st.divider()
            st.subheader("🔎 Ver / Editar Proveedor")

            prov_nombre = st.selectbox("Selecciona Proveedor", df['nombre'].tolist())
            row_sel     = df[df['nombre'] == prov_nombre].iloc[0]
            prov_id_sel = int(row_sel['id'])
            indice_sel  = calcular_indice(row_sel)

            m1, m2, m3 = st.columns(3)
            with m1: st.metric("📊 Índice Documental", f"{indice_sel}%")
            with m2:
                docs_ok = sum(1 for k in DOCUMENTOS if int(row_sel.get(k) or 0) == 1)
                st.metric("Docs Entregados", f"{docs_ok} / {TOTAL_DOCS}")
            with m3:
                if   indice_sel >= 80: st.success("🟢 COMPLETO")
                elif indice_sel >= 50: st.warning("🟡 EN PROCESO")
                else:                  st.error("🔴 CRÍTICO")

            with st.expander("📋 Ver información completa"):
                ic1, ic2 = st.columns(2)
                with ic1:
                    st.write(f"**NIT / Cédula:** {row_sel.get('nit_cedula','')}")
                    st.write(f"**Tipo:** {row_sel.get('tipo_bien_servicio','')}")
                    st.write(f"**Actividad CIUU:** {row_sel.get('tipo_actividad','')}")
                    st.write(f"**Dirección:** {row_sel.get('direccion_ciudad','')}")
                    st.write(f"**Teléfono:** {row_sel.get('telefono','')}")
                    st.write(f"**Contacto:** {row_sel.get('contacto','')}")
                    st.write(f"**Correo:** {row_sel.get('correo','')}")
                    st.write(f"**Fecha Vinculación:** {row_sel.get('fecha_vinculacion','')}")
                    st.write(f"**Última Actualización:** {row_sel.get('ultima_actualizacion','')}")
                    st.write(f"**Próxima Actualización:** {row_sel.get('proxima_actualizacion','')}")
                with ic2:
                    riesgo  = row_sel.get('eval_inicial_riesgo', '')
                    color_r = "🔴" if riesgo == "ALTO" else "🟡" if riesgo == "MEDIO" else "🟢" if riesgo == "BAJO" else "⚪"
                    st.write(f"**Nivel de Riesgo:** {color_r} {riesgo}")
                    st.write(f"**Fecha Eval.:** {row_sel.get('eval_inicial_fecha','')}")
                    st.write(f"**Reevaluación:** {row_sel.get('reevaluacion','')}")
                    st.write(f"**Control Visitas:** {row_sel.get('control_visitas','')}")
                    st.write(f"**Retroalimentación:** {row_sel.get('envio_retroalimentacion','')}")
                    st.write(f"**Otros Docs:** {row_sel.get('otros_documentos','')}")

            with st.expander("✏️ Editar documentos y subir PDFs"):
                st.markdown(f"### Editando: **{prov_nombre}**")
                st.caption(
                    "Marca ✔ los documentos entregados · "
                    "Sube el PDF de cada uno · "
                    "Descarga o elimina versiones anteriores"
                )

                doc_edit = {}
                for key, label in DOCUMENTOS.items():
                    current = bool(int(row_sel.get(key) or 0))
                    doc_edit[key] = widget_documento_pdf(
                        db=db,
                        proveedor_id=prov_id_sel,
                        doc_key=key,
                        doc_label=label,
                        checked=current,
                        form_key_prefix=f"e{prov_id_sel}",
                    )

                st.divider()
                st.markdown("**📅 Actualizar fechas y datos generales**")

                with st.form(f"form_meta_{prov_id_sel}"):
                    # NIT y tipo actividad
                    fn1, fn2 = st.columns(2)
                    with fn1:
                        nit_e = st.text_input("NIT / Cédula",
                                              value=str(row_sel.get('nit_cedula', '') or ''))
                    with fn2:
                        ciuu_idx_actual = get_ciuu_index(str(row_sel.get('tipo_actividad', '') or ''))
                        tipo_act_e = st.selectbox(
                            "Tipo de Actividad (CIUU)",
                            options=CIUU_LABELS,
                            index=ciuu_idx_actual,
                        )

                    ff1, ff2, ff3 = st.columns(3)
                    with ff1: vinc_e     = st.text_input("Fecha Vinculación",     value=str(row_sel.get('fecha_vinculacion', '') or ''))
                    with ff2: ult_act_e  = st.text_input("Última Actualización",  value=str(row_sel.get('ultima_actualizacion', '') or ''))
                    with ff3: prox_act_e = st.text_input("Próxima Actualización", value=str(row_sel.get('proxima_actualizacion', '') or ''))

                    fr1, fr2 = st.columns(2)
                    with fr1:
                        riesgo_ops = ["", "BAJO", "MEDIO", "ALTO"]
                        riesgo_act = str(row_sel.get('eval_inicial_riesgo', '') or '')
                        riesgo_e   = st.selectbox("Nivel de Riesgo", riesgo_ops,
                                                   index=riesgo_ops.index(riesgo_act)
                                                   if riesgo_act in riesgo_ops else 0)
                        eval_fech_e = st.text_input("Fecha Eval. Inicial",
                                                     value=str(row_sel.get('eval_inicial_fecha', '') or ''))
                    with fr2:
                        reeval_e  = st.text_input("Reevaluación",    value=str(row_sel.get('reevaluacion', '') or ''))
                        visitas_e = st.text_input("Control Visitas", value=str(row_sel.get('control_visitas', '') or ''))

                    retro_e = st.text_input("Retroalimentación", value=str(row_sel.get('envio_retroalimentacion', '') or ''))
                    otros_e = st.text_area("Otros Documentos",   value=str(row_sel.get('otros_documentos', '') or ''))

                    if st.form_submit_button("💾 Guardar Cambios", type="primary"):
                        datos_edit = {
                            'nit_cedula':         nit_e,
                            'nombre':             row_sel['nombre'],
                            'tipo_bien_servicio': row_sel.get('tipo_bien_servicio', ''),
                            'tipo_actividad':     tipo_act_e if tipo_act_e != CIUU_LABELS[0] else '',
                            'direccion_ciudad':   row_sel.get('direccion_ciudad', ''),
                            'telefono':           row_sel.get('telefono', ''),
                            'contacto':           row_sel.get('contacto', ''),
                            'correo':             row_sel.get('correo', ''),
                            **doc_edit,
                            'fecha_vinculacion':     vinc_e,
                            'ultima_actualizacion':  ult_act_e,
                            'proxima_actualizacion': prox_act_e,
                            'eval_inicial_fecha':    eval_fech_e,
                            'eval_inicial_riesgo':   riesgo_e,
                            'reevaluacion':          reeval_e,
                            'control_visitas':       visitas_e,
                            'envio_retroalimentacion': retro_e,
                            'otros_documentos':      otros_e,
                        }
                        if db.actualizar_proveedor(prov_id_sel, datos_edit):
                            st.success("✅ Proveedor actualizado correctamente")
                            st.rerun()

            st.divider()
            if 'confirmar_eliminar' not in st.session_state:
                st.session_state.confirmar_eliminar = None

            if st.button("🗑️ Eliminar este proveedor", type="secondary"):
                st.session_state.confirmar_eliminar = prov_id_sel

            if st.session_state.get('confirmar_eliminar') == prov_id_sel:
                st.warning(
                    f"⚠️ ¿Seguro que deseas eliminar **{prov_nombre}**? "
                    f"Se eliminarán también todos sus PDFs guardados."
                )
                bc1, bc2 = st.columns(2)
                with bc1:
                    if st.button("✅ Sí, eliminar", type="primary"):
                        db.eliminar_proveedor(prov_id_sel)
                        st.session_state.confirmar_eliminar = None
                        st.success("Proveedor eliminado.")
                        st.rerun()
                with bc2:
                    if st.button("❌ Cancelar"):
                        st.session_state.confirmar_eliminar = None
                        st.rerun()

    # ===========================================================
    # TAB 3 – REPORTES Y EXPORTACIÓN
    # ===========================================================
    with tab3:
        st.header("📊 Reportes y Exportación")
        df = db.obtener_proveedores()

        if df.empty:
            st.info("No hay datos para reportar aún.")
        else:
            indices_rep = [calcular_indice(r) for _, r in df.iterrows()]
            prom_rep    = sum(indices_rep) / len(indices_rep) if indices_rep else 0

            rc1, rc2, rc3, rc4 = st.columns(4)
            with rc1: st.metric("Total",          len(df))
            with rc2: st.metric("Índice Promedio", f"{prom_rep:.1f}%")
            with rc3: st.metric("🔴 Críticos",     sum(1 for i in indices_rep if i < 50))
            with rc4: st.metric("🟢 Completos",    sum(1 for i in indices_rep if i >= 80))

            st.divider()

            df_chart = df[['nombre']].copy()
            df_chart['Índice'] = indices_rep
            df_chart = df_chart.sort_values('Índice', ascending=True)
            fig1 = px.bar(df_chart, x='Índice', y='nombre', orientation='h',
                          title="📊 Índice de Cumplimiento por Proveedor",
                          color='Índice',
                          color_continuous_scale=['#FF4B4B', '#FFC300', '#28B463'],
                          range_color=[0, 100],
                          labels={'Índice': '% Cumplimiento', 'nombre': 'Proveedor'})
            fig1.add_vline(x=80, line_dash="dash", line_color="green",  annotation_text="Meta 80%")
            fig1.add_vline(x=50, line_dash="dash", line_color="orange", annotation_text="Mínimo 50%")
            fig1.update_layout(height=max(300, len(df) * 40))
            st.plotly_chart(fig1, use_container_width=True)

            st.divider()

            # Gráfico por tipo de actividad
            df_act_chart = df.copy()
            df_act_chart['_indice'] = indices_rep
            df_act_chart['_actividad'] = df_act_chart['tipo_actividad'].fillna('').replace('', 'Sin Actividad')
            # Tomar solo el código CIUU (primeros caracteres antes del " - ")
            df_act_chart['_actividad_corta'] = df_act_chart['_actividad'].apply(
                lambda x: x.split(' - ')[0] if ' - ' in x else x
            )
            act_grp = df_act_chart.groupby('_actividad_corta')['_indice'].mean().reset_index()
            act_grp.columns = ['Actividad CIUU', '% Promedio']
            act_grp = act_grp.sort_values('% Promedio', ascending=True)
            fig_act = px.bar(act_grp, x='% Promedio', y='Actividad CIUU', orientation='h',
                             title="📂 % Promedio de Cumplimiento por Tipo de Actividad CIUU",
                             color='% Promedio',
                             color_continuous_scale=['#FF4B4B', '#FFC300', '#28B463'],
                             range_color=[0, 100],
                             labels={'% Promedio': '% Cumplimiento Promedio'})
            fig_act.add_vline(x=80, line_dash="dash", line_color="green", annotation_text="Meta 80%")
            fig_act.update_layout(height=max(300, len(act_grp) * 45))
            st.plotly_chart(fig_act, use_container_width=True)

            st.divider()
            doc_pcts = [
                round(int(df[k].sum()) / len(df) * 100, 1) if k in df.columns else 0
                for k in DOCUMENTOS
            ]
            fig2 = px.bar(
                pd.DataFrame({'Documento': list(DOCUMENTOS.values()), '% Entrega': doc_pcts}),
                x='% Entrega', y='Documento', orientation='h',
                title="📄 % de Entrega por Tipo de Documento",
                color='% Entrega',
                color_continuous_scale=['#FF4B4B', '#FFC300', '#28B463'],
                range_color=[0, 100],
            )
            fig2.add_vline(x=80, line_dash="dash", line_color="green")
            fig2.update_layout(height=500)
            st.plotly_chart(fig2, use_container_width=True)

            st.divider()
            if 'eval_inicial_riesgo' in df.columns:
                rc = df['eval_inicial_riesgo'].replace('', 'SIN EVALUAR').value_counts().reset_index()
                rc.columns = ['Riesgo', 'Cantidad']
                fig3 = px.pie(rc, values='Cantidad', names='Riesgo',
                              title="🎯 Distribución de Nivel de Riesgo",
                              color='Riesgo',
                              color_discrete_map={'ALTO': '#FF4B4B', 'MEDIO': '#FFC300',
                                                  'BAJO': '#28B463', 'SIN EVALUAR': '#AAAAAA'})
                st.plotly_chart(fig3, use_container_width=True)

            st.divider()
            st.subheader("📥 Exportar a Excel")
            st.markdown(
                "6 hojas: **Directorio** · **Documentos y Cumplimiento** · "
                "**Evaluaciones** · **Informe Ejecutivo** · **Trazabilidad Actualizaciones** · "
                "**Análisis por Actividad CIUU**"
            )
            if st.button("⚙️ Generar Reporte Excel", type="primary"):
                with st.spinner("Generando..."):
                    excel_data = generar_excel_proveedores(df)
                st.download_button(
                    label="📥 Descargar Reporte",
                    data=excel_data,
                    file_name=f"Gestión_Proveedores_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                st.success("✅ Listo para descargar")


if __name__ == "__main__":
    main()
